import io
from typing import Dict, Any, List, Optional
from datetime import datetime
from .base import StructuredOutputProcessor, ContentParser
import logging
import re

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl库未安装，Excel功能将不可用")


class ExcelProcessor(StructuredOutputProcessor):
    """Excel文档输出处理器"""

    def __init__(self, **config):
        super().__init__(**config)
        if not OPENPYXL_AVAILABLE:
            raise ImportError("需要安装openpyxl库才能使用Excel功能")

        self.auto_width = config.get('auto_width', True)
        self.freeze_panes = config.get('freeze_panes', True)
        self.apply_styles = config.get('apply_styles', True)
        self.sheet_name = config.get('sheet_name', '数据')

    def get_content_type(self) -> str:
        """获取内容类型"""
        return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def get_file_extension(self) -> str:
        """获取文件扩展名"""
        return '.xlsx'

    def create_workbook(self) -> openpyxl.Workbook:
        """创建工作簿"""
        wb = openpyxl.Workbook()
        # 删除默认工作表
        if wb.active:
            wb.remove(wb.active)
        return wb

    def apply_header_style(self, worksheet, row: int):
        """应用表头样式"""
        if not self.apply_styles:
            return

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in worksheet[row]:
            if cell.value:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

    def apply_border(self, worksheet, start_row: int, end_row: int, start_col: int, end_col: int):
        """应用边框"""
        if not self.apply_styles:
            return

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                worksheet.cell(row=row, column=col).border = thin_border

    def auto_adjust_column_width(self, worksheet):
        """自动调整列宽"""
        if not self.auto_width:
            return

        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # 最大宽度限制为50
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def process_table(self, data: List[Dict[str, Any]], headers: Optional[List[str]] = None) -> bytes:
        """处理表格数据"""
        if not data:
            raise ValueError("数据不能为空")

        wb = self.create_workbook()
        ws = wb.create_sheet(title=self.sheet_name)

        # 如果没有提供headers，从第一行数据中提取
        if headers is None and data:
            headers = list(data[0].keys())

        if not headers:
            raise ValueError("无法确定表格标题")

        # 写入表头
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # 应用表头样式
        self.apply_header_style(ws, 1)

        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                # 处理不同数据类型
                if isinstance(value, (dict, list)):
                    value = str(value)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 应用边框
        if data:
            self.apply_border(ws, 1, len(data) + 1, 1, len(headers))

        # 冻结首行
        if self.freeze_panes:
            ws.freeze_panes = 'A2'

        # 自动调整列宽
        self.auto_adjust_column_width(ws)

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def process_list(self, items: List[str], ordered: bool = False) -> bytes:
        """处理列表数据"""
        if not items:
            raise ValueError("列表不能为空")

        wb = self.create_workbook()
        ws = wb.create_sheet(title=self.sheet_name)

        # 表头
        if ordered:
            headers = ['序号', '内容']
            ws.cell(row=1, column=1, value='序号')
            ws.cell(row=1, column=2, value='内容')
        else:
            headers = ['内容']
            ws.cell(row=1, column=1, value='内容')

        # 应用表头样式
        self.apply_header_style(ws, 1)

        # 写入数据
        for idx, item in enumerate(items):
            if ordered:
                ws.cell(row=idx + 2, column=1, value=idx + 1)
                ws.cell(row=idx + 2, column=2, value=item)
            else:
                ws.cell(row=idx + 2, column=1, value=item)

        # 应用边框
        end_col = 2 if ordered else 1
        self.apply_border(ws, 1, len(items) + 1, 1, end_col)

        # 冻结首行
        if self.freeze_panes:
            ws.freeze_panes = 'A2'

        # 自动调整列宽
        self.auto_adjust_column_width(ws)

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def process_key_value(self, data: Dict[str, Any]) -> bytes:
        """处理键值对数据"""
        if not data:
            raise ValueError("数据不能为空")

        wb = self.create_workbook()
        ws = wb.create_sheet(title=self.sheet_name)

        # 表头
        ws.cell(row=1, column=1, value='键')
        ws.cell(row=1, column=2, value='值')

        # 应用表头样式
        self.apply_header_style(ws, 1)

        # 写入数据
        for idx, (key, value) in enumerate(data.items(), 2):
            ws.cell(row=idx, column=1, value=str(key))
            # 处理复杂数据类型
            if isinstance(value, (dict, list)):
                value = str(value)
            ws.cell(row=idx, column=2, value=value)

        # 应用边框
        self.apply_border(ws, 1, len(data) + 1, 1, 2)

        # 冻结首行
        if self.freeze_panes:
            ws.freeze_panes = 'A2'

        # 自动调整列宽
        self.auto_adjust_column_width(ws)

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def process(self, content: str, metadata: Optional[Dict[str, Any]] = None) -> bytes:
        """处理内容并生成Excel文档"""
        if not self.validate_content(content):
            raise ValueError("内容不能为空")

        # 解析内容
        sections = ContentParser.parse_ai_response(content)
        tables = ContentParser.extract_tables_from_markdown(content)

        wb = self.create_workbook()

        # 如果找到表格，优先处理表格
        if tables:
            for i, table in enumerate(tables):
                sheet_name = f"表格{i+1}" if len(tables) > 1 else "表格"
                ws = wb.create_sheet(title=sheet_name)

                headers = table.get('headers', [])
                data = table.get('data', [])

                if headers and data:
                    # 写入表头
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header)

                    # 应用表头样式
                    self.apply_header_style(ws, 1)

                    # 写入数据
                    for row_idx, row_data in enumerate(data, 2):
                        for col_idx, cell_value in enumerate(row_data, 1):
                            ws.cell(row=row_idx, column=col_idx, value=cell_value)

                    # 应用边框
                    self.apply_border(ws, 1, len(data) + 1, 1, len(headers))

                    # 冻结首行
                    if self.freeze_panes:
                        ws.freeze_panes = 'A2'

                    # 自动调整列宽
                    self.auto_adjust_column_width(ws)

        # 如果没有表格，创建文本内容工作表
        if not tables:
            ws = wb.create_sheet(title="文档内容")

            # 准备元数据
            doc_metadata = self.prepare_metadata(metadata)

            # 写入元数据
            current_row = 1
            if doc_metadata:
                ws.cell(row=current_row, column=1, value="文档信息")
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
                current_row += 1

                for key, value in doc_metadata.items():
                    if value:
                        ws.cell(row=current_row, column=1, value=key)
                        ws.cell(row=current_row, column=2, value=str(value))
                        current_row += 1

                current_row += 1  # 空行

            # 写入内容sections
            ws.cell(row=current_row, column=1, value="文档内容")
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
            current_row += 1

            for section in sections:
                section_type = section.get('type', 'paragraph')
                content_text = section.get('content', '')

                if section_type == 'heading':
                    level = section.get('level', 1)
                    font_size = max(12, 16 - level)
                    ws.cell(row=current_row, column=1, value=content_text)
                    ws.cell(row=current_row, column=1).font = Font(bold=True, size=font_size)
                elif section_type == 'list':
                    items = section.get('items', [])
                    ordered = section.get('ordered', False)
                    for i, item in enumerate(items):
                        prefix = f"{i+1}. " if ordered else "• "
                        ws.cell(row=current_row, column=1, value=f"{prefix}{item}")
                        current_row += 1
                    current_row -= 1  # 调整，因为循环会多加一行
                else:
                    ws.cell(row=current_row, column=1, value=content_text)

                current_row += 1

            # 自动调整列宽
            self.auto_adjust_column_width(ws)

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def process_multi_sheet_data(self, sheets_data: Dict[str, List[Dict[str, Any]]],
                                metadata: Optional[Dict[str, Any]] = None) -> bytes:
        """处理多个工作表的数据"""
        if not sheets_data:
            raise ValueError("工作表数据不能为空")

        wb = self.create_workbook()

        for sheet_name, data in sheets_data.items():
            if not data:
                continue

            ws = wb.create_sheet(title=sheet_name)

            # 获取表头
            headers = list(data[0].keys()) if data else []

            if headers:
                # 写入表头
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

                # 应用表头样式
                self.apply_header_style(ws, 1)

                # 写入数据
                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        value = row_data.get(header, '')
                        if isinstance(value, (dict, list)):
                            value = str(value)
                        ws.cell(row=row_idx, column=col_idx, value=value)

                # 应用边框
                self.apply_border(ws, 1, len(data) + 1, 1, len(headers))

                # 冻结首行
                if self.freeze_panes:
                    ws.freeze_panes = 'A2'

                # 自动调整列宽
                self.auto_adjust_column_width(ws)

        # 如果有元数据，创建信息工作表
        if metadata:
            doc_metadata = self.prepare_metadata(metadata)
            info_ws = wb.create_sheet(title="文档信息", index=0)

            current_row = 1
            for key, value in doc_metadata.items():
                if value:
                    info_ws.cell(row=current_row, column=1, value=key)
                    info_ws.cell(row=current_row, column=2, value=str(value))
                    current_row += 1

            # 应用样式
            self.apply_header_style(info_ws, 1)
            self.auto_adjust_column_width(info_ws)

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def process_chat_conversation(self, conversation: List[Dict[str, str]],
                                metadata: Optional[Dict[str, Any]] = None) -> bytes:
        """处理聊天对话为Excel格式"""
        if not conversation:
            raise ValueError("对话内容不能为空")

        wb = self.create_workbook()
        ws = wb.create_sheet(title="对话记录")

        # 表头
        headers = ['序号', '角色', '时间', '内容']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # 应用表头样式
        self.apply_header_style(ws, 1)

        # 写入对话数据
        for idx, message in enumerate(conversation, 2):
            role = message.get('role', 'user')
            content = message.get('content', '')
            timestamp = message.get('timestamp', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

            ws.cell(row=idx, column=1, value=idx - 1)  # 序号
            ws.cell(row=idx, column=2, value='用户' if role == 'user' else 'AI助手')
            ws.cell(row=idx, column=3, value=timestamp)
            ws.cell(row=idx, column=4, value=content)

        # 应用边框
        self.apply_border(ws, 1, len(conversation) + 1, 1, 4)

        # 冻结首行
        if self.freeze_panes:
            ws.freeze_panes = 'A2'

        # 自动调整列宽
        self.auto_adjust_column_width(ws)

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


def create_excel_processor(**config) -> ExcelProcessor:
    """创建Excel处理器实例"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("需要安装openpyxl库才能使用Excel功能")
    return ExcelProcessor(**config)
